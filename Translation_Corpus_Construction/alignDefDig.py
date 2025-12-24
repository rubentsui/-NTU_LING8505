import sys, os
import regex as re
from datetime import datetime, timedelta
from pathlib import Path, PurePath
from math import ceil
from random import seed as seed
import numpy as np
import sqlite3
from tqdm import tqdm 
import torch
from sentence_splitter import SentenceSplitter, split_text_into_sentences
#from nltk import word_tokenize
import unicodedata
import pysbd
import opencc

import ujson
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows


from dp_utils import make_alignment_types, read_alignments, \
    read_in_embeddings, make_doc_embedding, vecalign, yield_overlaps

from score import score_multiple, log_final_scores

from sentence_transformers import SentenceTransformer, models, util

s2tw = opencc.OpenCC('s2tw.json')


start_time = datetime.now()
'''
if os.name == 'nt':
    d = 1
elif os.name == 'posix':
    if torch.cuda.is_available():
        d = 0
    elif torch.mps.is_available():
        d = 1
'''
# Model loading moved to __main__ to allow import
model = None
dev = None


#%%

def encodeVectors(ss, model, dev):
    '''
    Input:
        ss: lit of strings 
    Output:
        pytorch tensors
    '''
    vecs = model.encode(ss, show_progress_bar=False, convert_to_numpy=False, normalize_embeddings=True, device=dev)
    return vecs

def print_alignments(alignments, scores=None, file=sys.stdout):
    if scores is not None:
        for (x, y), s in zip(alignments, scores):
            print('%s:%s:%.6f' % (x, y, s), file=file)
    else:
        for x, y in alignments:
            print('%s:%s' % (x, y), file=file)


def file_open(filepath):
    #Function to allowing opening files based on file extension
    if filepath.endswith('.gz'):
        return gzip.open(filepath, 'rt', encoding='utf8')
    elif filepath.endswith('.bz2'):
        return bz2.open(filepath, 'rt', encoding='utf8')
    elif filepath.endswith('.xz'):
        return lzma.open(filepath, 'rt', encoding='utf8')
    else:
        return open(filepath, 'r', encoding='utf8')

def getLines(fin):
    '''
    Retrive lines from a file or (later) sqlite3 database
    '''
    lines = file_open(fin).readlines()
    return [s.strip() for s in lines if s.strip() != '']

def getSentIndex(lines):
    """
    dictionary look-up:
        keys = sentence or overlapped sentences
        value = index
    """
    sent2line = dict()
    for ii, line in enumerate(lines):
        if line.strip() in sent2line:
            raise Exception('got multiple embeddings for the same line')
        sent2line[line.strip()] = ii
    return sent2line

def getOverlaps(lines, num_overlaps):
    output = set()
    for out_line in yield_overlaps(lines, num_overlaps):
        output.add(out_line)

    # for reproducibility
    output = list(output)
    output.sort()
    return output

def normalizeText(text):
    text = text.replace("\xad", '')  # remove Unicode soft hyphen
    return unicodedata.normalize("NFKC", text) # remove Unicode , among others

# Sentence tokenizer

# regex to identify Chinese sentence boundaries
#regex_zh_sent_delim = re.compile(r"([。！？；][」』”〕》〗】)）\]]?)")
#regex_zh_sent_delim = re.compile(r"([。？；][」』”〕》〗】)）\]]?)")
#regex_zh_sent_delim = re.compile(r'(?P<quotation_mark>([。？！…]{1,2})[」』〕》〗】\])”’"\'）])')
#regex_zh_sent_delim = re.compile(r"[。！？]")
regex_zh_sent_delim = re.compile(r"([。？！…][」』”’\'\"〕》〗】)）\]]{0,3})")

def normalizeTextZh(text):
    text = text.replace("\xad", '')  # remove Unicode 
    #text = text.replace("!", "！").replace(";", "；")
    return unicodedata.normalize("NFKD", text) # remove Unicode , among others

def sentencizeZh(s):
    '''
    turn long string s into a list of sentences
    '''
    s = normalizeTextZh(s)
    s = s.replace(',','，').replace(';','；').replace("!", "！").replace(":", "：").replace("?", "？")
    ss = regex_zh_sent_delim.sub(r"\1\n", s).split("\n")
    return [s.strip() for s in ss if s.strip() != '']


def sentencize(s, lang='en'):
    if lang in ['zh', 'ja']:
        return sentencizeZh(s)
    else: # lang in ['en', 'es', 'fr', 'de', 'it', etc. ]
        splitter = SentenceSplitter(language=lang)
        sentseg = pysbd.Segmenter(language=lang, clean=False)
        s = normalizeText(s)
        ss = splitter.split(text=s)
        #ss = sentseg.segment(s)
        return [s.strip() for s in ss if s.strip() != '']

def convertChinesePunctuations(txt):
    '''
    Convert “”‘’ to, respeectively 「」『』 
    '''
    punctHans2Hant = {'“”‘’': '「」『』'}
    for k in punctHans2Hant:
        v = punctHans2Hant[k]
        for ps, pt in zip(k, v):
            txt = txt.replace(ps, pt)
    return txt
    

def align(sS, sT, alignment_max_size=4):
     
    # make runs consistent
    seed(42)
    np.random.seed(42)

    # source
    overlapsS = getOverlaps(sS, alignment_max_size)  # create "overlapped" sentences
    s2idxS = getSentIndex(overlapsS)                 # create "sentence-to-index" lookup table
    embedS = encodeVectors(overlapsS, model, dev)    # encode a list of sentences 
    src_line_embeddings = torch.vstack(embedS).cpu().numpy()   # turns a list of sentences into a tensor object
    # target
    overlapsT = getOverlaps(sT, alignment_max_size)
    s2idxT = getSentIndex(overlapsT)
    embedT = encodeVectors(overlapsT, model, dev)
    overlapsS = getOverlaps(sS, alignment_max_size)
    tgt_line_embeddings = torch.vstack(embedT).cpu().numpy()
    
    #print(f"src_line_embeddings has shape: [{src_line_embeddings.shape}]")
    #print(f"tgt_line_embeddings has shape: [{tgt_line_embeddings.shape}]")
    #sys.exit(0)

    width_over2 = ceil(alignment_max_size / 2.0) + 5

    test_alignments = []
    stack_list = []
    
    #src_lines = open(finS, 'rt', encoding="utf-8").readlines()
    vecs0 = make_doc_embedding(s2idxS, src_line_embeddings, sS, alignment_max_size)

    #tgt_lines = open(finT, 'rt', encoding="utf-8").readlines()
    vecs1 = make_doc_embedding(s2idxT, tgt_line_embeddings, sT, alignment_max_size)

    final_alignment_types = make_alignment_types(alignment_max_size)

    stack = vecalign(vecs0=vecs0,
                     vecs1=vecs1,
                     final_alignment_types=final_alignment_types,
                     del_percentile_frac=0.2,
                     width_over2=width_over2,
                     max_size_full_dp=300,
                     costs_sample_size=20000,
                     num_samps_for_norm=100)

    # write final alignments to fk\ile
    #print_alignments(stack[0]['final_alignments'], stack[0]['alignment_scores'])
    #test_alignments.append(stack[0]['final_alignments'])
    #stack_list.append(stack)
    
    alignments = stack[0]['final_alignments']
    scores     = stack[0]['alignment_scores']

    aligned_sentences = []
    if scores is not None:
        for (idxS, idxT), score in zip(alignments, scores):
            sbS  = [] # sentence block - source
            for i in idxS:
                sbS.append(sS[i])            
            sbT  = [] # sentence block - target
            for i in idxT:
                sbT.append(sT[i])
            
            #aligned_sentences.append(f"{score:.5f}\t{idxS}\t{' '.join(sbS)}\t{idxT}\t{' '.join(sbT)}")            
            aligned_sentences.append([score, idxS, ' '.join(sbS), idxT, ' '.join(sbT)])      
    return aligned_sentences
#%%

def createExcel(fin):
    
    """ fin = plain text aligned text
    """
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Select the active sheet
    ws = wb.active
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 65
    
    data = open(fin, 'r', encoding='utf-8').readlines()

    df = pd.DataFrame([x.split('\t') for x in data], columns=['cosdist', 'cols_s', langS, 'cols_t',  langT])

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    # Set cell alignment
    alignment = Alignment(horizontal='general',
                          vertical='top',
                          wrap_text=True)
    cnt = len(data)
    for row in ws[f'A1:F{cnt+10}']:
        for cell in row:
            cell.alignment = alignment

    # Save the workbook
    base = Path(fin).stem
    fon_xlsx = Path(fin).parent / f'{base}.{langS}-{langT}.xlsx'
    wb.save(fon_xlsx)

#%%

if __name__ == '__main__':
    
    # Initialize model and device
    m = int(sys.argv[1])
    d = 0
    dev = ['cuda', 'mps', 'cpu'][d]
    #Model we want to use for bitext mining. LaBSE achieves state-of-the-art performance
    model_name = ['ibm-granite/granite-embedding-278m-multilingual', 'LaBSE', 'Alibaba-NLP/gte-multilingual-base', 'paraphrase-multilingual-MiniLM-L12-v2'][m]
    model_name_short = ['ibm-granite-278m', 'LaBSE', 'alibaba-gte', 'MiniLM-L12-v2'][m]

    if 'model' not in globals() or model is None:
        print(f"Now running bitext mining with transformer model [{model_name}] on device [{dev}]...", flush=True)
        model = SentenceTransformer(model_name, device=dev)
        print(f"Finished loading model: {model_name}.", flush=True)
    else:
        print(f"Model [{model_name}] already loaded", flush=True)

    end_time = datetime.now() - start_time
    print(f"Model-loading time: {end_time.seconds} secs", flush=True)

    alignment_max_size = 7
    print(f"alignment_max_size = {alignment_max_size}")
    
    ###########################################################
    # Step 1 Use chapter separator?
    # Step 2 Convert to Traditional Chinese?
    ###########################################################
    USE_REGEX_CHAPTER_SEPARATOR = False # True #   False
    CONVERT_ZHS_TO_ZHT = False # True # False # True

    ###########################################################
    # Step 3 Choose language pair (translation direction)
    ###########################################################
    langS = ['zh', 'en', 'fr', 'zh', 'en', 'es', 'zh', 'ja', 'es'][0]
    langT = ['zh', 'en', 'fr', 'it', 'zhs'][1]

    out_langS, out_langT = langS, langT

    ###########################################################
    # Step 5 Choose input file folder 
    ###########################################################
    #base_folder = '/Users/rubentsui/NLP/DefenseDigestVol52'
    base_folder = '.'

    ######################################################################
    # regex for dividing text into chunks (chapter, book, section, etc.)

    #base_fn = '國防譯粹 第52卷第04期_訓練革新_263392'
    #base_fn = '國防譯粹 第52卷第02期_領導思維_445840'
    base_fn = sys.argv[2]
    
    fin = f"{base_folder}/{base_fn}.populated.json"
    DEFDIG = ujson.load(open(fin, 'r', encoding='utf-8'))

    for idx, a in enumerate(DEFDIG):

        #if idx > 1: break

        fon = f"{base_folder}/{base_fn}.vecalign.n{alignment_max_size}.{out_langS}-{out_langT}.{model_name_short}.txt"

        print(a['chinese_title'])
        print(a['english_source_title'])

        S = a['similarity']
        largest_similarity_idx = S.index(max(S))
        print("Candidate:", a['english_source_candidate_titles'][largest_similarity_idx])
        print(f"  with similarity: {S[largest_similarity_idx]:.4f}")
        if S[largest_similarity_idx] < 0.6:
            print("Skipping... not similar enough")
            print('-'*20)
            continue

        txtS = a['chinese_paragraphs']
        chS = ['\n'.join(txtS)]

        txtT = a['english_paragraphs']
        chT = ['\n'.join(txtT[largest_similarity_idx])]
        
        ch_cnt = 0
        for cS, cT in zip(chS, chT):
            
            ch_cnt += 1
            print(f"processing segment [{ch_cnt}]...", flush=True)
    
            # Source    
            pS = cS.strip().split("\n")
            pS = [s.strip() for s in pS if s.strip()!='']
            sS = []
            for p in pS:
                sS.extend(sentencize(p, lang=langS))
            sS = [s.strip() for s in sS if s.strip()!='']
            ## convert source from simplified Chinese to traditional Chinese
    
            # Target
            pT = cT.strip().split("\n")
            pT = [s.strip() for s in pT if s.strip()!='']
            sT = []
            for p in pT:
                sT.extend(sentencize(p, lang=langT))
            sT = [s.strip() for s in sT if s.strip()!='']
            ## convert target from simplified Chinese to traditional Chinese
            if CONVERT_ZHS_TO_ZHT and langT == 'zh':
                #sT = [s2tw.convert(s).replace('“','「').replace('”','」') for s in sT]
                sT = [convertChinesePunctuations(s2tw.convert(s)) for s in sT]

            with open(fon, "a", encoding="utf-8", newline="\n") as fo:
                #for score, idxE, e, idxZ, z in align(sE, sZ, alignment_max_size=alignment_max_size):
                fo.write(f"{1:.4f}\t[0]\t{a['chinese_title']}\t[0]\t{a['english_source_title']}\n")
                for score, idxS, ss, idxT, tt in align(sS, sT, alignment_max_size=alignment_max_size):
                    #fo.write(f"{base}\t{score:.4f}\t{idxS}\t{ss}\t{idxT}\t{tt}\n")
                    fo.write(f"{score:.4f}\t{idxS}\t{ss}\t{idxT}\t{tt}\n")
                    fo.flush()

    print('-'*25)
    fon_xlsx = fon
    print("Creating Excel file...")
    createExcel(fon_xlsx)
    print('='*25)



#%%
